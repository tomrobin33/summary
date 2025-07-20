from openpyxl import load_workbook

def parse_excel(path):
    wb = load_workbook(path, data_only=True)
    result = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        result.append({"sheet": sheet, "data": data})
    return result 